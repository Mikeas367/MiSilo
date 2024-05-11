import tkinter as tk
from pathlib import Path
from tkinter import messagebox
from datetime import date
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

try:  # creo un archivo de txt para guardar el peso del silo
    with open('peso_silo.txt', 'r') as file:
        contenido_silo = int(file.read())  # leo el contenido del txt
except (FileNotFoundError, ValueError):  # esto hace que si la línea está vacía no salte error
    contenido_silo = 0  # si no existía un archivo inicio la variable


def guardar_en_excel(cantidad, tipo_de_la_bolsa, peso_de_la_bolsa, cliente):
    # Comprobar si el archivo "Bolsas.xlsx" existe
    file_path = Path('Bolsas.xlsx')
    if file_path.is_file():
        # Cargar el archivo existente en modo de lectura y escritura
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        # Creo un nuevo archivo si no existe
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Fecha'
        sheet['B1'] = 'Cantidad'
        sheet['C1'] = 'Kilos de la Bolsa'
        sheet['D1'] = 'Tipo de Bolsa'
        sheet['E1'] = 'Cliente'

        # Obtener la fecha actual en formato dd/mm/yyyy
    fecha_actual = date.today().strftime('%d/%m/%Y')

    # Buscar la primera fila vacía a partir de la fila 2
    first_empty_row = 2
    while sheet.cell(row=first_empty_row, column=1).value is not None:
        first_empty_row += 1

    # Escribir los nuevos datos en la primera fila vacía
    sheet.cell(row=first_empty_row, column=1).value = fecha_actual
    sheet.cell(row=first_empty_row, column=2).value = cantidad
    sheet.cell(row=first_empty_row, column=3).value = str(peso_de_la_bolsa) + "kg"
    sheet.cell(row=first_empty_row, column=4).value = tipo_de_la_bolsa
    sheet.cell(row=first_empty_row, column=5).value = cliente
    # Guardar el libro de trabajo en el archivo
    workbook.save(file_path)
    messagebox.showinfo("Confirmación", "Datos guardados en el archivo Excel.")


def opciones_de_bolsas(vector, ventana):
    # Opciones de bolsa
    opciones = vector
    ventana_aplicar = ventana
    tipo_bolsa_var = tk.StringVar(ventana_aplicar)
    tipo_bolsa_var.set(opciones[0])  # Establecer la opción por defecto

    tipo_bolsa_label = tk.Label(ventana_aplicar, text="Seleccionar tipo de bolsa:", fg="white", bg="green",
                                font=("Arial", 12))
    tipo_bolsa_label.pack(pady=5)

    tipo_bolsa_option_menu = tk.OptionMenu(ventana_aplicar, tipo_bolsa_var, *opciones)
    tipo_bolsa_option_menu.config(bg="yellow", fg="black", font=("Arial", 10))
    tipo_bolsa_option_menu.pack()
    return tipo_bolsa_var


def ventana_bolsas():
    ventana_bolsas = tk.Tk()
    ventana_bolsas.title("Bolsas")
    ventana_bolsas.geometry("800x500")  # Establecer el tamaño de la ventana (ancho x alto)
    ventana_bolsas.configure(bg="green")

    titulo_label = tk.Label(ventana_bolsas, text="Ingresar el cliente", fg="white", bg="green",
                            font=("Arial", 14, "bold"))
    titulo_label.pack(pady=10)

    # Cuadro para ingresar contenido
    cliente = tk.Entry(ventana_bolsas)
    cliente.pack()
    titulo_label = tk.Label(ventana_bolsas, text="Ingresar cantidad de bolsas", fg="white", bg="green",
                            font=("Arial", 14, "bold"))
    titulo_label.pack(pady=10)

    # Cuadro para ingresar cantidad de bolsas
    cantidad_bolsas = tk.Entry(ventana_bolsas)
    cantidad_bolsas.pack()

    # Opciones de bolsa
    tipos_de_bolsas = ["ENTERO", "MOLIDO", "PARTIDO"]
    kilos_de_bolsas = ["20", "30", "35"]
    seleccion_de_bolsa_var = opciones_de_bolsas(tipos_de_bolsas, ventana_bolsas)
    seleccion_de_kilo_var = opciones_de_bolsas(kilos_de_bolsas, ventana_bolsas)

    def obtener_seleccion():
        # obtengo los datos que se ingresaron por pantalla
        global contenido_silo

        seleccion_bolsa = seleccion_de_bolsa_var.get()
        seleccion_kilo = seleccion_de_kilo_var.get()
        cantidad_bolsas_ingresada = int(cantidad_bolsas.get())
        cliente_ingresado = cliente.get()
        peso_restar = cantidad_bolsas_ingresada * int(seleccion_kilo)
        if peso_restar >= contenido_silo:
            messagebox.showinfo("peso", f"No se pueden hacer estas bolsas te quedan {contenido_silo} kg en el silo")
        else:
            print(f"se restan de tu silo:{peso_restar}")
            contenido_silo -= peso_restar
            guardar_en_excel(cantidad_bolsas_ingresada, seleccion_bolsa, seleccion_kilo, cliente_ingresado)
            guardar_contenido()
            print("cliente ingresado: ", cliente_ingresado)
            print(f"cantidad de bolsas ingresadas: {cantidad_bolsas_ingresada}")
            print("Tipo de bolsa seleccionado:", seleccion_bolsa)
            print("Kilo de bolsa seleccionado:", seleccion_kilo)

    boton_mostrar_seleccion = tk.Button(ventana_bolsas, text="Aceptar", command=obtener_seleccion)
    boton_mostrar_seleccion.pack(pady=10)
    ventana_bolsas.mainloop()


def guardar_contenido():
    global contenido_silo
    with open('peso_silo.txt', 'w') as file:
        file.write(str(contenido_silo))


def agregar_contenido():
    global contenido_silo, entrada
    # Verificar si el peso actual alcanza el límite máximo
    if contenido_silo >= 60000:
        messagebox.showinfo("Límite alcanzado", "El silo ha alcanzado su límite máximo de capacidad.")
    else:
        # Obtener la cantidad ingresada por el usuario
        contenido_nuevo = entrada.get()
        if not contenido_nuevo.isdigit():
            messagebox.showinfo("Error", "Debes de ingresar números")
            return
        contenido_nuevo2 = int(contenido_nuevo)
        '''''''''
        creo contenido_nuevo2 porque si me ingresan letras quiero que me muestre el mensaje de error y si no son letras lo transformo a entero
        '''
        if contenido_silo + contenido_nuevo2 > 60000:  # verifico que con la cantidad nueva a ingresar y lo que ya tenía no supere el límite
            messagebox.showinfo("Límite alcanzado",
                                "No es posible agregar el contenido. Superaría el límite máximo de capacidad.")
        else:
            contenido_silo += contenido_nuevo2
            guardar_contenido()
            messagebox.showinfo("Contenido agregado",
                                f"Se agregaron {contenido_nuevo2} kg de contenido al silo.\n\nPeso actual del silo: "
                                f"{contenido_silo} kg")
            entrada.delete(0, tk.END)  # Limpiar el campo de entrada


def ver_silo():
    messagebox.showinfo("Mi Silo", f"El peso actual del silo es: {contenido_silo} kg")


def ventana():
    global entrada  # pongo como variable global entrada para poder usarla en la función agregar contenido
    ventana_principal = tk.Tk()  # creo la ventana principal
    ventana_principal.title("Mi Silo")  # Título de la ventana
    ventana_principal.geometry("800x500")  # tamaño de la ventana (ancho x alto)
    ventana_principal.configure(bg="green")  # color de fondo de la ventana principal
    # Título del cuadro para ingresar contenido al silo
    titulo_label = tk.Label(ventana_principal, text="Ingresar maíz a mi silo", fg="white", bg="green",
                            font=("Arial", 14, "bold"))
    titulo_label.pack(pady=10)
    # cuadro para ingresar contenido
    entrada = tk.Entry(ventana_principal)
    entrada.pack()
    # BOTONES
    # botón para agregar contenido
    agregar_boton = tk.Button(ventana_principal, text="Agregar contenido a mi silo", command=agregar_contenido,
                              bg="yellow", width=25, height=3)
    agregar_boton.pack(pady=10)
    # botón para ver el contenido del silo
    ver_silo_boton = tk.Button(ventana_principal, text="Ver mi silo", command=ver_silo,
                               bg="yellow", width=25, height=3)
    ver_silo_boton.pack(pady=10)
    # botón para abrir la ventana de fabricación de las bolsas
    abrir_ventana_bolsas_boton = tk.Button(ventana_principal, text="Hacer bolsas", command=ventana_bolsas,
                                           bg="yellow", width=25, height=3)
    abrir_ventana_bolsas_boton.pack(pady=10)
    ventana_principal.mainloop()  # loop de la ventana principal


def main():
    ventana()


if __name__ == '__main__':
    main()
