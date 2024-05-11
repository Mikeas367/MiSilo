from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from pathlib import Path
from datetime import date
import pandas as pd
import tkinter as tk

# Variables
peso_actual = 0
altura_rectangulo = 0
peso_total_resta = 0
datos_excel = pd.read_excel('Bolsas.xlsx')


def guardar_en_excel(cantidad, tipo_de_la_bolsa, peso_de_la_bolsa, cliente):
    # Comprobar si el archivo "Bolsas.xlsx" existe
    file_path = Path('Bolsas.xlsx')
    if file_path.is_file():
        # Cargar el archivo existente en modo de lectura y escritura
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        # Creo un nuevo archivo si no existe
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Fecha'
        sheet['B1'] = 'Tipo de bolsa'
        sheet['C1'] = 'Kilos de la Bolsa'
        sheet['D1'] = 'Cantidad'
        sheet['E1'] = 'Cliente'

    # Obtener la fecha actual en formato dd/mm/yyyy
    fecha_actual = date.today().strftime('%d/%m/%Y')

    # Buscar la primera fila vacía a partir de la fila 2
    first_empty_row = 2
    while sheet.cell(row=first_empty_row, column=1).value is not None:
        first_empty_row += 1

    # Escribir los nuevos datos en la primera fila vacía
    sheet.cell(row=first_empty_row, column=1).value = fecha_actual
    sheet.cell(row=first_empty_row, column=2).value = tipo_de_la_bolsa
    sheet.cell(row=first_empty_row, column=3).value = str(peso_de_la_bolsa) + "kg"
    sheet.cell(row=first_empty_row, column=4).value = cantidad
    sheet.cell(row=first_empty_row, column=5).value = cliente

    # Guardar el libro de trabajo en el archivo
    workbook.save(file_path)


# Función para agregar contenido al silo
def agregar_contenido():
    global peso_actual

    # Verificar si el peso actual alcanza el límite máximo
    if peso_actual >= 60000:
        messagebox.showinfo("Límite alcanzado", "El silo ha alcanzado su límite máximo de capacidad.")
    else:
        # Obtener la cantidad ingresada por el usuario
        contenido_nuevo = int(entrada.get())

        # Verificar si el contenido nuevo supera el límite
        if peso_actual + contenido_nuevo > 60000:
            messagebox.showinfo("Límite alcanzado",
                                "No es posible agregar el contenido. Superaría el límite máximo de capacidad.")
        else:
            peso_actual += contenido_nuevo
            if peso_actual >= 1:
                guardar_en_excel(1, "Silo", contenido_nuevo, "Silo")
            messagebox.showinfo("Contenido agregado",
                                f"Se agregaron {contenido_nuevo} kg de contenido al silo.\n\nPeso actual del silo: "
                                f"{peso_actual} kg")
            entrada.delete(0, tk.END)  # Limpiar el campo de entrada
    actualizar_rectangulo()


def ver_silo():
    messagebox.showinfo("Mi Silo", f"El peso actual del silo es: {peso_actual} kg")


# Función para abrir la ventana de hacer bolsas

def abrir_ventana_bolsas():
    ventana_bolsas = tk.Toplevel()
    ventana_bolsas.title("Bolsas")
    ventana_bolsas.geometry("500x400")  # Establecer el tamaño de la ventana (ancho x alto)
    ventana_bolsas.configure(bg="lightblue")  # Establecer el color de fondo de la ventana de bolsas

    # Variable para almacenar la cantidad de bolsas
    cantidad_bolsas = tk.StringVar()

    # Variable para almacenar el peso de la bolsa seleccionada
    peso_bolsa = tk.IntVar()
    tipo_bolsa = tk.StringVar()
    nombre_cliente = tk.StringVar()

    # =SUMAR.SI(B:B, "Entero", D:D)
    # Función para hacer las bolsas
    def hacer_bolsas():
        global peso_actual
        global peso_total_resta
        nombre = nombre_cliente.get()
        # Obtener la cantidad de bolsas ingresada
        cantidad = int(cantidad_bolsas.get())
        # Obtener el peso de la bolsa seleccionada
        peso_de_la_bolsa = peso_bolsa.get()
        tipo_de_la_bolsa = tipo_bolsa.get()
        # EXCEL DE LAS BOLSAS
        guardar_en_excel(cantidad, tipo_de_la_bolsa, peso_de_la_bolsa, nombre)
        # Calcular el peso total a restar del silo
        peso_restar = cantidad * peso_de_la_bolsa

        # Verificar si el peso a restar supera el peso actual del silo
        if peso_restar > peso_actual:
            messagebox.showinfo("Error", "No hay suficiente contenido en el silo para hacer esa cantidad de bolsas.")
        else:
            peso_actual -= peso_restar
            peso_total_resta += peso_restar
            print("Peso Total Resta = ", peso_total_resta)
            messagebox.showinfo("Bolsas hechas",
                                f"Se han sacado {peso_restar} kg\n\nSe han hecho {cantidad} bolsas de "
                                f"{peso_de_la_bolsa} "
                                f"kg cada una.\n\nPeso actual del silo: {peso_actual} kg")
            ventana_bolsas.destroy()  # Cerrar la ventana de bolsas
        print("Nombre del cliente:", nombre)
        actualizar_rectangulo()

    # Cliente
    nombre_label = tk.Label(ventana_bolsas, text="Nombre del cliente:", bg="lightblue")
    nombre_label.pack()
    nombre_entry = tk.Entry(ventana_bolsas, textvariable=nombre_cliente)
    nombre_entry.pack()

    # Etiqueta y campo de entrada para la cantidad de bolsas
    cantidad_label = tk.Label(ventana_bolsas, text="Cantidad de bolsas:", bg="lightblue")
    cantidad_label.pack()
    cantidad_entry = tk.Entry(ventana_bolsas, textvariable=cantidad_bolsas)
    cantidad_entry.pack()

    # Frame para los botones de peso de bolsa
    botones_frame = tk.Frame(ventana_bolsas, bg="lightblue")
    botones_frame.pack(pady=10)

    # Botones para seleccionar el peso de las bolsas
    bolsa20_boton = tk.Button(botones_frame, text="20 kg", command=lambda: peso_bolsa.set(20), bg="yellow")
    bolsa20_boton.pack(side="left", padx=10)

    bolsa30_boton = tk.Button(botones_frame, text="30 kg", command=lambda: peso_bolsa.set(30), bg="yellow")
    bolsa30_boton.pack(side="left", padx=10)

    bolsa35_boton = tk.Button(botones_frame, text="35 kg", command=lambda: peso_bolsa.set(35), bg="yellow")
    bolsa35_boton.pack(side="left", padx=10)

    # Botones para Los Tipos de Bolsa
    bolsa_entera = tk.Button(botones_frame, text="Bolsa Entero", command=lambda: tipo_bolsa.set("Entero"), bg="yellow")
    bolsa_entera.pack(side="left", padx=10)

    bolsa_partido = tk.Button(botones_frame, text="Bolsa Partido", command=lambda: tipo_bolsa.set("Partido"), bg="ye"
                                                                                                                 "llow")
    bolsa_partido.pack(side="left", padx=10)

    bolsa_molido = tk.Button(botones_frame, text="Bolsa Molido", command=lambda: tipo_bolsa.set("Molido"), bg="yellow")
    bolsa_molido.pack(side="left", padx=10)

    # Botón para hacer las bolsas
    hacer_bolsas_boton = tk.Button(ventana_bolsas, text="Hacer bolsas", command=hacer_bolsas)
    hacer_bolsas_boton.pack()
    ventana_bolsas.grab_set()  # Bloquear la interacción con la ventana principal
    ventana_bolsas.protocol("WM_DELETE_WINDOW",
                            ventana_bolsas.destroy)  # Configurar el comportamiento de cierre de la ventana
    ventana_bolsas.mainloop()


# Crear una instancia de la ventana principal
ventana = tk.Tk()
ventana.title("Mi Silo")
ventana.geometry("800x500")  # Establecer el tamaño de la ventana (ancho x alto)
ventana.configure(bg="green")  # Establecer el color de fondo de la ventana principal

# Etiqueta para el título de ingresar contenido al silo
titulo_label = tk.Label(ventana, text="Ingresar maíz a mi silo", fg="white", bg="green", font=("Arial", 14, "bold"))
titulo_label.pack(pady=10)

# Campo de entrada para la cantidad de contenido a agregar
entrada = tk.Entry(ventana)
entrada.pack()

# Botón para agregar contenido al silo
agregar_boton = tk.Button(ventana, text="Agregar Contenido al silo", command=agregar_contenido, bg="yellow",
                          width=25, height=3)
agregar_boton.pack(pady=10)

# Botón para ver el contenido actual del silo
ver_silo_boton = tk.Button(ventana, text="Ver Mi Silo", command=ver_silo, bg="yellow", width=25, height=3)
ver_silo_boton.pack()

# Botón para abrir la ventana de hacer bolsas
hacer_bolsas_boton = tk.Button(ventana, text="Hacer Bolsas", command=abrir_ventana_bolsas, bg="yellow", width=25,
                               height=3)
hacer_bolsas_boton.pack(pady=10)
# IMAGEN SILO

lienzo_rectangulo = tk.Canvas(ventana, width=200, height=400, bg="white")
lienzo_rectangulo.place(x=550, y=50)


# actualizar imagen del silo
def actualizar_rectangulo():
    # Calcular la altura del rectángulo en función del peso_actual
    altura_rectangulo = peso_actual / 60000 * 400

    # Borrar cualquier dibujo anterior en el lienzo
    lienzo_rectangulo.delete("all")

    # Dibujar el rectángulo gris en el lienzo
    lienzo_rectangulo.create_rectangle(0, 0, 250, 400, fill="lightgray")

    # Dibujar el rectángulo de llenado en función de la altura
    lienzo_rectangulo.create_rectangle(0, 400 - altura_rectangulo, 250, 400, fill="yellow")

    # Dibujar el graduado
    for i in range(0, 60001, 10000):
        y = 400 - (i / 60000 * 400)
        lienzo_rectangulo.create_line(0, y, 250, y, fill="black")
        lienzo_rectangulo.create_text(105, y, anchor=tk.N, text=str(i))


def cerrar_ventana():
    # Guardar el peso actual del silo y la altura del rectángulo en el archivo
    with open("peso_silo.txt", "w") as file:
        file.write(f"{peso_actual}\n{altura_rectangulo}")
    ventana.destroy()


# Cargar el peso actual del silo desde el archivo de guardado

# Cargar el peso actual y la altura del rectángulo desde el archivo de guardado
try:
    with open("peso_silo.txt", "r") as file:
        peso_actual = int(file.readline())
        altura_line = file.readline().strip()
        altura_rectangulo = int(altura_line) if altura_line else 0
except FileNotFoundError:
    # Si el archivo no existe, se crea con los valores iniciales de 0
    with open("peso_silo.txt", "w") as file:
        file.write("0\n0")

# Actualizar el rectángulo al iniciar la aplicación
actualizar_rectangulo()
# Configurar el comportamiento de cierre de la ventana principal
ventana.protocol("WM_DELETE_WINDOW", cerrar_ventana)

# Iniciar el bucle principal de la ventana principal
ventana.mainloop()
